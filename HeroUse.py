import pandas as pd
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.formatting.rule import DataBar, FormatObject,Rule,DataBarRule
from openpyxl import load_workbook
#计算出场率
#

#常量区域
race_types=(1011,1111,1022)#比赛类型
sum_count=0
max_pvp_level=0
level_type=''
def get_winrate(heroid,pvp_level,race_type,data):
    '''通过英雄id，获取该段位、该游戏比赛类型下的胜率'''
    
    use_data=data[(data.hero==heroid) & (data[level_type]==pvp_level) & (data.subtype==race_type)]
    
    win_number=0
    lose_number=0
    
    if len(use_data[use_data.is_win==1])==1:
        win_number=use_data[use_data.is_win==1].times.iloc[0]
    if len(use_data[use_data.is_win==0])==1:
        lose_number=use_data[use_data.is_win==0].times.iloc[0]
       
    try:
        return win_number/(lose_number+win_number)
        
    except ZeroDivisionError:
        

        return 0
def set_times_sum(data):
    '''计算所有比赛类型、段位的比赛场次总和'''
    global level_type
    resoult={}
    
    for race_type in race_types:
        resoult[race_type]={}
        for level in range(1,max_pvp_level+1):
            resoult[race_type][level]=data[(getattr(data,"subtype")==race_type) & (getattr(data,level_type)==level)]['times'].sum()/8
            print("总和计算完成"+str(resoult[race_type][level]))

    return resoult,len(race_types)*max_pvp_level
def get_userate(heroid,pvp_level,race_type,data,times_sum_dict):
    global level_type
    '''获取英雄使用率'''
    try:
        if times_sum_dict[race_type][pvp_level]==0:
            return 0
        return (data[(data.subtype==race_type) & (data[level_type]==pvp_level)&(data.hero==heroid)]['times'].sum())/times_sum_dict[race_type][pvp_level]
    except ZeroDivisionError:

        return 0

def build_excel(file_path,version,bulider):
    with pd.ExcelWriter(file_path,mode='w',engine='openpyxl') as writer:
            pd.DataFrame({'version':version,
            'bulider':bulider},index=["信息"]).to_excel(writer,sheet_name='Infomation_Page')
def set_excel_format(file_path):
    #百分比，微软雅黑 11号 有边框，数据条格式
    font=Font(
        name='微软雅黑',
        size=10
    )
    border= Border(
        left=Side(border_style=None,color='FF000000'),
        right=Side(border_style=None,color='FF000000'),
        top=Side(border_style=None,color='FF000000'),
        bottom=Side(border_style=None,color='FF000000')
    )

    #数据条
    
    
    rule=DataBarRule(start_type='percent', start_value='0', end_type='percent', end_value='100',color="00FF0000", showValue="None", minLength=None, maxLength=None)
    
    #打开工作簿，进行格式设定
    excel=load_workbook(file_path)
    for sheetname in race_types:
        sheetname=str(sheetname)
        table=excel[sheetname]
        table.freeze_panes='D2'
        for row in range(2,table.max_row+1):
            for column in range(2,table.max_column+1):
                
                table.cell(row,column).font=font
                table.cell(row,column).border=border
                table.cell(row,column).number_format='0.00%'

        table.conditional_formatting.add('B2:AI69',rule)
        
    excel.save(file_path)
def handle_file(read_file,save_file,th,user_data): 
    global max_pvp_level
    global level_type
    max_pvp_level=user_data['level']
    level_type=user_data['level_type']

    save_file+=user_data['file_name']
    hero_property_data=pd.read_csv("Data/hero_property.csv",encoding='gb2312')
    
    try:
        hero_use_data=pd.read_csv(read_file,encoding ='gb2312',keep_default_na=False)
        
    except FileNotFoundError:
        return
    
    hero_ids=hero_property_data["hero_id"].tolist()
    
    build_excel(save_file,'1.0.1','xxxx')
    times_sum_dict,sum_count=set_times_sum(hero_use_data)
    
    #循环
    count=0
    for race_type in race_types:
        to_excel_dict={}
        for level in range(1,user_data['level']+1):
            print(str(race_type),str(level)+"正在计算中....")
            if level==1:
                to_excel_dict["英雄名称"]={}
                to_excel_dict["英雄品质"]={} 
            to_excel_dict[user_data['col_name']+str(level)+"使用率"]={}
            to_excel_dict[user_data['col_name']+str(level)+"胜率"]={}
            
            for heroid in hero_ids:
                #print(hero_property_data[hero_property_data.hero_id==heroid]["hero_name"].iloc[0])
                to_excel_dict["英雄名称"][heroid]=hero_property_data[hero_property_data.hero_id==heroid]["hero_name"].iloc[0]
                to_excel_dict["英雄品质"][heroid]=hero_property_data[hero_property_data.hero_id==heroid]["hero_color"].iloc[0]            
                to_excel_dict[user_data['col_name']+str(level)+"使用率"][heroid]=get_userate(heroid,level,race_type,hero_use_data,times_sum_dict)  
                to_excel_dict[user_data['col_name']+str(level)+"胜率"][heroid]=get_winrate(heroid,level,race_type,hero_use_data)
            print(str(race_type),str(level)+"计算完成！")
            count+=1
            th.trigger.emit(count/sum_count*100)
        #转换为dataframe
        #print(to_excel_dict["英雄名称"][heroid])
        to_excel_df=pd.DataFrame(to_excel_dict)
        
        #写入excel
        with pd.ExcelWriter(save_file,mode='a',engine='openpyxl') as writer:
            to_excel_df.to_excel(writer,sheet_name=str(race_type))
    #设定格式
    set_excel_format(save_file)

        
            
    
    